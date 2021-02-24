from django.shortcuts import render
# 引入Projectdata的类
from projectmanage.models import Projectdata
# 引入JsonResponse模块
from django.http import JsonResponse
# 导入json模块
import json
# 导入Q查询 或者查询
from django.db.models import Q
# 导入uuid类
import uuid
# 导入哈希库
import hashlib
# 导入Setting
from django.conf import settings
# 导入os
import os
# 引入处理Excel模块
import openpyxl
# Create your views here.

def query_projectdata(request):
    """查询项目信息"""
    # 接收传递过来的查询条件--- axios默认是json --- 字典类型（'inputstr'）-- data['inputstr']
    data = json.loads(request.body.decode('utf-8'))
    try:
        # 使用ORM获取所有项目信息 并把对象转为字典格式
        obj_projectdata = Projectdata.objects.filter(Q(time__icontains=data['inputstr']) |  Q(projectName__icontains=data['inputstr'])).values()
        # 把外层的容器转为List
        projectdata = list(obj_projectdata)
        # 返回
        return JsonResponse({'code':1, 'data':projectdata})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取项目信息出现异常，具体错误：" + str(e)})


def add_projectdata(request):
    """添加学生到数据库"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 添加到数据库
        obj_projectdata = Projectdata(time=data['time'],projectName=data['projectName'],man=data['man'],
                              days=data['days'],manDay=data['manDay'],
                              bugNumber= data['bugNumber'], bugRate=data['bugRate'])
        # 执行添加
        obj_projectdata.save()
        # 使用ORM获取所有项目信息 并把对象转为字典格式
        obj_projectdata = Projectdata.objects.all().values()
        # 把外层的容器转为List
        projectdata = list(obj_projectdata)
        # 返回
        return JsonResponse({'code': 1, 'data': projectdata})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "添加到数据库出现异常，具体原因：" + str(e)})

def update_projectdata(request):
    """修改学生到数据库"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 查找到要修改的项目信息
        obj_projectdata = Projectdata.objects.get(projectName=data['projectName'])
        # 依次修改
        obj_projectdata.time = data['time']
        obj_projectdata.projectName = data['projectName']
        obj_projectdata.man = data['man']
        obj_projectdata.days = data['days']
        obj_projectdata.manDay = data['manDay']
        obj_projectdata.bugNumber = data['bugNumber']
        obj_projectdata.bugRate = data['bugRate']
        # 保存
        obj_projectdata.save()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_projectdatas = Projectdata.objects.all().values()
        # 把外层的容器转为List
        projectdatas = list(obj_projectdatas)
        # 返回
        return JsonResponse({'code': 1, 'data': projectdatas})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "修改保存到数据库出现异常，具体原因：" + str(e)})

def delete_projectdata(request):
    """删除项目"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 查找到要修改的项目信息
        obj_projectdata = Projectdata.objects.get(projectName=data['projectName'])
        # 删除
        obj_projectdata.delete()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_projectdatas = Projectdata.objects.all().values()
        # 把外层的容器转为List
        projectdatas = list(obj_projectdatas)
        # 返回
        return JsonResponse({'code': 1, 'data': projectdatas})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "删除项目出现异常，具体原因：" + str(e)})

def delete_projectdatas(request):
    """批量删除多条项目"""
    # 接收前端传递过来的值
    data = json.loads(request.body.decode("utf-8"))
    try:
        # 遍历传递的集合
        for one_project in data['projects']:
            # 查询当前记录
            obj_projectdata = Projectdata.objects.get(projectName=one_project['projectName'])
            # 删除
            obj_projectdata.delete()
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_projectdatas = Projectdata.objects.all().values()
        # 把外层的容器转为List
        projectdatas = list(obj_projectdatas)
        # 返回
        return JsonResponse({'code': 1, 'data': projectdatas})
    except Exception as e:
        return JsonResponse({'code':0 , 'msg': "删除项目出现异常，具体原因：" + str(e)})

def is_exists_projectname(request):
    """判断项目名称是否存在"""
    # 接收传递过来的项目名称
    data = json.loads(request.body.decode('utf-8'))
    # 进行校验
    try:
        obj_projectdata = Projectdata.objects.filter(projectName=data['projectName'])
        if obj_projectdata.count() == 0:
            return JsonResponse({'code': 1, 'exists': False})
        else:
            return JsonResponse({'code': 1, 'exists': True})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg':"校验项目名称失败，具体原因：" + str(e)})

def get_projectdata(request):
    """获取所有学生的信息"""
    try:
        # 使用ORM获取所有学生信息 并把对象转为字典格式
        obj_projectdata = Projectdata.objects.all().values()
        # 把外层的容器转为List
        students = list(obj_projectdata)
        # 返回
        return JsonResponse({'code':1, 'data':students})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取学生信息出现异常，具体错误：" + str(e)})

def import_projectdata_execl(request):
    """从Excel批量导入学生信息"""
    # ========1. 接收Excel文件存储到Media文件夹 =======
    rev_file = request.FILES.get('execl')
    # 判断，是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': 'Execl文件不存在！'})
    # 获得一个唯一的名字： uuid +hash
    new_name = get_random_str()
    # 准备写入的URL
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1])
    # 开始写入到本次磁盘
    try:
        f = open(file_path, 'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        # 要关闭
        f.close()
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': str(e)})

    #====== 2. 读取存储在Media文件夹的数据  =====
    ex_students = read_excel_dict(file_path)

    # ====3. 把读取的数据存储到数据库 =====
    # 定义几个变量： success:  error: errors
    success = 0
    error = 0
    error_projectNames = []

    # 开始遍历
    for one_student in ex_students:
        print(one_student)
        try:
            obj_student = Projectdata.objects.create(time=one_student['time'], projectName=one_student['projectName'], days=one_student['days'],
                                                      man=one_student['man'],manDay=one_student['manDay'],bugNumber=one_student['bugNumber'],bugRate=one_student['bugRate'])
            print(111111)
            print(obj_student)
            # 计数
            success += 1
        except:
            # 如果失败了
            error += 1
            error_projectNames.append(one_student['projectName'])


    # 4. 返回--导入信息（成功：5，失败：4--（projectNames））,所有学生
    obj_projectdatas = Projectdata.objects.all().values()
    projectdatas = list(obj_projectdatas)
    return JsonResponse({'code':1, 'success':success,'error':error,'errors':error_projectNames, 'data':projectdatas})



def export_projectdata_execl(request):
    """导出数据到excel"""
    # 获取所有的学生信息
    obj_students = Projectdata.objects.all().values()
    # 转为List
    projectdata = list(obj_students)
    # 准备名称
    excel_name = get_random_str() + ".xlsx"
    # 准备写入的路劲
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到Excel
    write_to_excel(projectdata, path)
    # 返回
    return JsonResponse({'code':1, 'name':excel_name })

def write_to_excel(data:list, path:str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = '项目BUG情况统计'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k,v in enumerate(keys):
            sheet.cell(row=index + 1, column=k+ 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)

def get_random_str():
    #获取uuid的随机数
    uuid_val = uuid.uuid4()
    #获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    #获取md5实例
    md5 = hashlib.md5()
    #拿取uuid的md5摘要
    md5.update(uuid_str)
    #返回固定长度的字符串
    return md5.hexdigest()

def read_excel_dict(path:str):
    """读取Excel数据，存储为字典 --- [{},{},{},]"""
    # 实例化一个wrokbook
    workbook = openpyxl.load_workbook(path)
    # 实例化一个sheet
    sheet = workbook['项目BUG情况统计']
    # 定义一个变量存储最终的数据--[]
    projectdata = []
    # 准备key
    keys = ['time','projectName','days','man','manDay','bugNumber','bugRate']
    # 遍历
    for row in sheet.rows:
        # 定义一个临时的字典
        temp_dict = {}
        # 组合值和key
        for index,cell in enumerate(row):
            # 组和
            temp_dict[keys[index]] = cell.value
        # 附加到list中
        projectdata.append(temp_dict)
    #返回
    return projectdata